# -*- coding: utf-8 -*-
"""
Created on Mon Mar  7 19:41:38 2022

@author: Test-PC
"""
import random
import gc
import os
import numpy as np
import matplotlib.pyplot as plt
from netpyne import specs, sim
from joblib import Parallel, delayed
import multiprocessing
import itertools
from openpyxl import load_workbook
import gc


# Network parameters

netParams = specs.NetParams()  # object of class NetParams to store the network parameters

L23e = {'secs': {}}

L23e['secs']['soma'] = {'geom': {}, 'mechs': {}}
L23e['secs']['soma']['geom'] = {'diam': 96, 'L': 96, 'Ra': 100, 'cm': 1, 'nseg' : 1}
L23e['secs']['soma']['mechs']['pas'] = {'e':-70,'g':0.0001} 
L23e['secs']['soma']['mechs']['hh2'] = {'gnabar': 0.05, 'gkbar': 0.005, 'vtraub':-55 }
L23e['secs']['soma']['mechs']['im'] = {'gkbar':7e-5} 

netParams.cellParams['PYR'] = L23e


## Population parameters
netParams.popParams['S'] = {'cellType': 'PYR', 'numCells': 1}
netParams.popParams['M'] = {'cellType': 'PYR', 'numCells': 1}

## Synaptic mechanism parameters

###############################################################################  Synaptic mechanism parameters
# netParams.synMechParams['exc']   = {'mod': 'Exp2Syn', 'tau1': 1.0, 'tau2': 5.0, 'e': 0}  # excitatory synaptic mechanism
# netParams.synMechParams['AMPA']  = {'mod': 'Exp2Syn', 'tau1': 0.2, 'tau2': 1.7,    'e': 0,   'g': 0.2}     # AMPA synaptic mechanism
# netParams.synMechParams['NMDA']  = {'mod': 'Exp2Syn', 'tau1': 2,   'tau2': 26,     'e': 0,   'g': 0.03}    # NMDA synaptic mechanism
# netParams.synMechParams['GABAa'] = {'mod': 'Exp2Syn', 'tau1': 0.3, 'tau2': 2.5,    'e': -70, 'g': 0.5}     # GABAa synaptic mechanism   
# netParams.synMechParams['GABAb'] = {'mod': 'Exp2Syn', 'tau1': 45.2,'tau2': 175.16, 'e': -70, 'g': 0.05}    # GABAb synaptic mechanism
netParams.synMechParams['STD']   = {'mod': 'FDSExp2Syn'} #, 'tau_D1': 140,'d1': 0.7 , 'tau_D2' : 8200}                            # STD synaptic mechanism   

###############################################################################  Population parameters
        # f = 0.917 (1) < 0, 1e9 >    : facilitation
        # tau_F = 94 (ms) < 1e-9, 1e9 >
        # d1 = 0.416 (1) < 0, 1 >     : fast depression
        # tau_D1 = 380 (ms) < 1e-9, 1e9 >
        # d2 = 0.975 (1) < 0, 1 >     : slow depression
        # tau_D2 = 9200 (ms) < 1e-9, 1e9 >

###############################################################################
last_stim=110840
for t in (500-last_stim,1000 ):
    delay=last_stim+t
    Input_A='Plast_test'+ str(t)
    Target_A='Plast_test'+ str(t)+'->S'
    amplitude= 75 #np.random.uniform(1,1.3,1)
    netParams.stimSourceParams[Input_A] = {'type': 'IClamp',  'del': delay, 'dur': 0.1 , 'amp':amplitude}  
    netParams.stimTargetParams[Target_A] = {'source': Input_A, 'sec':'soma', 'loc': 0.5,'conds': {'pop':['S']}}        
###############################################################################
###############################################################################

STDPparams = {'STDPon':1, 'hebbwt': 0.001, 'antiwt':-0.001, 'wmax': 50, 'RLon': 0 ,'tauhebb': 10, 'RLwindhebb': 0, 'useRLexp': 0, 'softthresh': 0, 'verbose':0}
# Cell connectivity rules
netParams.connParams['S->M'] = {'preConds': {'pop': 'S'}, 'postConds': {'pop': ['M']},  #  S -> M
    'probability': 1, #'normal(0.8,0.1)',           # probability of connection
    'weight': 0.05, #'negexp(0.3)',              # synaptic weight
    'delay': 1, #'normal(20,8)',                 # transmission delay (ms)
    'sec': 'soma',            # section to connect to
    'loc': 1.0,                 # location of synapse
    'plast': {'mech': 'STDP', 'params': STDPparams},
    # 'synMech': ['AMPA','NMDA','STD']}           # target synaptic mechanism
    # 'synMech': ['AMPA','NMDA']}           # target synaptic mechanism
    # 'synMech': ['AMPA','STD']}           # target synaptic mechanism
    # 'synMech': ['NMDA','STD']}           # target synaptic mechanism
    # 'synMech': 'AMPA'}           # target synaptic mechanism
    'synMech': 'STD'}           # target synaptic mechanism
###############################################################################
# Simulation options
simConfig = specs.SimConfig()       # object of class SimConfig to store simulation configuration

simConfig.duration = 113*1e3        # Duration of the simulation, in ms  #72*1e3
simConfig.dt = 0.025                # Internal integration timestep to use
simConfig.verbose = True           # Show detailed messages
simConfig.saveCellConns = True
simConfig.recordTraces = {
                          'V_soma':  {'sec': 'soma', 'loc': 0.9, 'var':'v'},
                          'g_STD':   {'sec': 'soma', 'loc': 1.0, 'synMech': 'STD', 'var': 'g'}
                          }
simConfig.hParams = {'celsius': 37, 'v_init': -65.0, 'clamp_resist': 0.001}

simConfig.recordStep = 0.05            # Step size in ms to save data (eg. V traces, LFP, etc)
simConfig.filename = 'test'         # Set file output name
simConfig.savePickle = False        # Save params, network and sim output to pickle file

# simConfig.analysis['plotRaster'] = {'saveFig': True}                    # Plot a raster
simConfig.analysis['plotTraces'] = {'include': [1], 'saveFig': True}  # Plot recorded traces for this list of cells
# simConfig.analysis['plot2Dnet'] = {'saveFig': True}                     # plot 2D cell positions and connections
# simConfig.analysis['plotConn']={'includePre':'S','includePost':'M','feature':'weight', 'saveFig': True }
 
###############################################################################
###############################################################################

threshold=60.00      #0.63 for orginal setting

def Simulation(amp_pre, amp_post):
    simConfig.filename='imTBS-pre-'+str(amp_pre)+'post-'+str(amp_post)
    # COUNT=0

    start_time=1000
    for ttt in range (0,110000,15000):
        for tt in range(0,5000,200):
            for t in range(0,50,20):
                delay=start_time+t+tt+ttt    
    #             COUNT=COUNT+1
    #             print (delay)
    # print ('COUNT='+str(COUNT))
                Input_PreSy='Input_PreSy'+ str(ttt)+ str(tt)+str(t) 
                Input_PostSy='Input_PostSy'+ str(ttt)+ str(tt)+str(t)
                Target_PreSy='Input_PreSy'+ str(ttt)+str(tt)+str(t)+'->S'
                Target_PostSy='Input_PostSy'+ str(ttt)+str(tt)+str(t)+'->S'
    
                netParams.stimSourceParams[Input_PreSy] =  {'type': 'IClamp',  'del': delay, 'dur': 0.1 , 'amp':threshold*amp_pre/100} 
                netParams.stimSourceParams[Input_PostSy] = {'type': 'IClamp',  'del': delay, 'dur': 0.1 , 'amp':threshold*amp_post/100}
     
                netParams.stimTargetParams[Target_PreSy]  = {'source': Input_PreSy,  'sec':'soma', 'loc': 0.5,'conds': {'pop':['S']}} 
                netParams.stimTargetParams[Target_PostSy] = {'source': Input_PostSy, 'sec':'soma', 'loc': 0.5,'conds': {'pop':['M']}} 
    sim.createSimulateAnalyze(netParams = netParams, simConfig = simConfig)
    response = sim.allSimData.g_STD.cell_1
    SimDataTime=sim.simData.t
    return amp_pre, amp_post, SimDataTime, response

num_cores = multiprocessing.cpu_count()-1
amp_pre=range  (99,160,4)
amp_post=range (44,157,4)
amps = list(itertools.product(amp_pre, amp_post))

chunk_size = 2
chunks = [amps[x:x+chunk_size] for x in range(0, len(amps), chunk_size)]

for chunk in chunks:
    Data=[]
    ModelOutput=Parallel(n_jobs=num_cores)(delayed(Simulation)(amp[0],amp[1]) for amp in chunk)   
    for i in range (chunk_size):
        pre_a=  ModelOutput[i][0]
        post_a= ModelOutput[i][1]
        Time=   ModelOutput[i][2]
        Voltage=ModelOutput[i][3]
        filename='imTBS-pre-'+str(pre_a)+'post-'+str(post_a)

        voltage_cut=[]
        time_cut=[]
        for j in range  (2236800,2237000,1):
            time_cut.append(Time[j])
            voltage_cut.append(Voltage[j])
        conductivity=np.max(voltage_cut)
        
        plt.figure()
        y = voltage_cut
        x = time_cut
        plt.plot(x, y)
        plt.xlabel('Time (ms)')
        plt.ylabel('recorded conductivity')
        plt.title(filename)
        plt.savefig(filename+'.png')
    
   
        xlsx_filename = "Raw_Data.xlsx"
        workbook = load_workbook(xlsx_filename)
        worksheet = workbook['imTBS']
        Column=int((pre_a-99)/4+3)
        Row=int((post_a-40)/4+4)
        worksheet.cell(row=Row, column=Column).value = conductivity 
        workbook.save(xlsx_filename)
    gc.collect()

